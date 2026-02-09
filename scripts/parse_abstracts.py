"""
Parse abstract Excel files into gold standard JSON.
Uses label-based search instead of hardcoded row numbers.

Run: source backend/venv/bin/activate && python scripts/parse_abstracts.py
"""
import openpyxl
import os
import json
import datetime
import re
import sys


def safe_val(v):
    """Convert Excel values to JSON-safe types."""
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    return s if s else None


def safe_str(v):
    """Always return string or None."""
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m-%d')
    s = str(v).strip()
    return s if s else None


def find_label(ws, label, max_row=200, col_range=range(1, 17)):
    """Find the row and column of a cell containing the given label."""
    label_lower = label.lower()
    for row in range(1, min(max_row, ws.max_row + 1)):
        for col in col_range:
            val = ws.cell(row, col).value
            if val and label_lower in str(val).lower():
                return row, col
    return None, None


def get_value_near_label(ws, label, offset_col=2, offset_row=0, max_row=200, col_range=range(1, 17)):
    """Find a label and return the value offset from it."""
    row, col = find_label(ws, label, max_row, col_range)
    if row is None:
        return None
    target_row = row + offset_row
    target_col = col + offset_col
    if target_col < 1 or target_row < 1:
        return None
    return safe_val(ws.cell(target_row, target_col).value)


def get_value_right_of_label(ws, label, max_row=200):
    """Find a label and return the value in the next non-empty cell to its right."""
    row, col = find_label(ws, label, max_row)
    if row is None:
        return None
    # Scan right for a value
    for c in range(col + 1, min(col + 6, 17)):
        v = ws.cell(row, c).value
        if v is not None:
            return safe_val(v)
    return None


def parse_format_a(ws):
    """Parse Format A abstract using label-based search."""
    data = {}

    # ── Property Info ──
    data['property_name'] = get_value_right_of_label(ws, 'Property Name:')
    data['property_number'] = get_value_right_of_label(ws, 'Property #:')
    data['building_sf'] = get_value_right_of_label(ws, 'Building SF:')
    data['percentage_leased'] = get_value_right_of_label(ws, 'Percentage Leased:')

    # ── Tenant Info ──
    data['tenant_legal_name'] = get_value_right_of_label(ws, 'Legal/Tenant Name:')
    data['tenant_sq_feet'] = get_value_right_of_label(ws, 'SQ. Feet:')

    # Addresses: multi-row, need to find the label row then read below it
    billing_row, _ = find_label(ws, 'Billing Address:')
    if billing_row:
        parts = [safe_str(ws.cell(billing_row, 4).value),
                 safe_str(ws.cell(billing_row + 1, 4).value),
                 safe_str(ws.cell(billing_row + 2, 4).value)]
        data['billing_address'] = ' '.join(filter(None, parts)) or None

    premise_row, _ = find_label(ws, 'Premise Address:')
    if premise_row:
        parts = [safe_str(ws.cell(premise_row, 11).value),
                 safe_str(ws.cell(premise_row + 1, 11).value),
                 safe_str(ws.cell(premise_row + 2, 11).value)]
        data['premise_address'] = ' '.join(filter(None, parts)) or None

    data['tenant_dba'] = get_value_right_of_label(ws, 'Tenant DBA:')
    data['guarantor'] = get_value_right_of_label(ws, 'Guarantor:')

    # Phone/email: on same rows as DBA/Guarantor
    phone_row, phone_col = find_label(ws, 'Phone:')
    if phone_row:
        data['phone'] = safe_str(ws.cell(phone_row, phone_col + 2).value)
    else:
        data['phone'] = None

    email_row, email_col = find_label(ws, 'Email:')
    if email_row is None:
        email_row, email_col = find_label(ws, 'E-mail:')
    if email_row:
        data['email'] = safe_str(ws.cell(email_row, email_col + 2).value)
    else:
        data['email'] = None

    # ── Lease Info ──
    data['occupancy_date'] = get_value_right_of_label(ws, 'Occupancy Date:')
    data['term_months'] = get_value_right_of_label(ws, 'Term:')

    # Dates: find by their specific labels
    data['rent_commencement'] = get_value_right_of_label(ws, 'Rent Commencement:')
    data['lease_commencement'] = get_value_right_of_label(ws, 'Lease Commencement:')
    data['execution_date'] = get_value_right_of_label(ws, 'Execution Date:')
    data['lease_expiration'] = get_value_right_of_label(ws, 'Lease Expiration:')
    data['move_in_date'] = get_value_right_of_label(ws, 'Move In Date:')

    data['prepaid_rent'] = get_value_right_of_label(ws, 'Prepaid Rent:')
    data['lease_type'] = get_value_right_of_label(ws, 'Type:')
    data['security_deposit'] = get_value_right_of_label(ws, 'Security Deposit:')
    data['holdover'] = get_value_right_of_label(ws, 'Holdover:')

    # ── Description / Use ──
    use_row, _ = find_label(ws, 'Description/Use:')
    if use_row:
        parts = [safe_str(ws.cell(use_row + r, 4).value) for r in range(0, 3)]
        data['permitted_use'] = ' '.join(filter(None, parts)) or None
    else:
        data['permitted_use'] = None

    data['exclusives'] = get_value_right_of_label(ws, 'Exclusives:')

    # ── Rent ──
    data['free_rent'] = get_value_right_of_label(ws, 'Free Rent:')
    data['late_charge'] = get_value_right_of_label(ws, 'Late Charge:')
    data['due_date'] = get_value_right_of_label(ws, 'Due Date:')

    # Rent schedule: find "Base Rent:" row, data starts 1 row below
    base_rent_row, _ = find_label(ws, 'Base Rent:')
    rent_schedule = []
    if base_rent_row:
        # The row with "Base Rent:" has column headers (Start Date, End Date, Annual, Monthly, SF/Yr)
        # Data rows start after that
        for r in range(base_rent_row + 1, base_rent_row + 15):
            start = safe_val(ws.cell(r, 4).value)
            end = safe_val(ws.cell(r, 6).value)
            annual = safe_val(ws.cell(r, 8).value)
            monthly = safe_val(ws.cell(r, 11).value)
            psf = safe_val(ws.cell(r, 13).value)
            label_val = safe_str(ws.cell(r, 2).value)

            # Skip if no dates or if we hit column headers
            if not start or not end:
                continue
            # Skip header row (contains strings like "Start Date")
            if isinstance(start, str) and not re.match(r'^\d{4}-\d{2}-\d{2}$', start):
                continue
            # Skip rows where annual/monthly are both 0 or None (empty schedule rows)
            if (annual is None or annual == 0) and (monthly is None or monthly == 0):
                # Still include free rent periods (annual=0, monthly=0 is valid)
                pass

            rent_schedule.append({
                'label': label_val,
                'start_date': start,
                'end_date': end,
                'annual': annual,
                'monthly': monthly,
                'psf': psf
            })

    data['rent_schedule'] = rent_schedule

    # Extract first full (non-prorated, non-zero) rent period for comparison
    if rent_schedule:
        for rs in rent_schedule:
            monthly = rs['monthly']
            annual = rs['annual']
            # Skip zero, None, prorated, or string values
            if not monthly or monthly == 0:
                continue
            if isinstance(monthly, str) or isinstance(annual, str):
                continue
            if rs.get('label') and 'prorat' in str(rs['label']).lower():
                continue
            data['base_rent_monthly'] = monthly
            data['base_rent_annual'] = annual
            data['rent_per_sf_annual'] = rs['psf']
            break

    # Rent comments
    rent_comments_row, _ = find_label(ws, 'Comments:', max_row=50)
    if rent_comments_row and base_rent_row and rent_comments_row > base_rent_row:
        parts = [safe_str(ws.cell(rent_comments_row + r, 4).value) for r in range(0, 3)]
        data['rent_comments'] = ' '.join(filter(None, parts)) or None
    else:
        data['rent_comments'] = None

    # ── Percentage Rent ──
    pct_row, _ = find_label(ws, 'Percent Rent:')
    data['percentage_rent'] = {}
    if pct_row:
        data['percentage_rent']['percent'] = safe_val(ws.cell(pct_row, 4).value)
        bp_val = get_value_right_of_label(ws, 'Breakpoint:')
        data['percentage_rent']['breakpoint'] = bp_val

    # ── Renewal / Options ──
    renewal_row, _ = find_label(ws, 'Renewal:')
    data['renewal_options'] = safe_str(ws.cell(renewal_row, 4).value) if renewal_row else None
    data['renewal_notice'] = safe_str(ws.cell(renewal_row, 11).value) if renewal_row else None

    expansion_row, _ = find_label(ws, 'Expansion:')
    data['expansion'] = safe_str(ws.cell(expansion_row, 4).value) if expansion_row else None

    termination_row, _ = find_label(ws, 'Termination:')
    data['termination'] = safe_str(ws.cell(termination_row, 4).value) if termination_row else None

    # ── Operating Expenses ──
    data['pro_rata_share'] = get_value_right_of_label(ws, 'Pro Rata Share:')

    cam_row, _ = find_label(ws, 'Common Area Maint')
    if cam_row:
        data['cam'] = safe_val(ws.cell(cam_row, 8).value)
        data['cam_comments'] = safe_str(ws.cell(cam_row, 11).value)
    else:
        data['cam'] = None
        data['cam_comments'] = None

    # CAM additional notes
    add_notes_row, _ = find_label(ws, 'Additional Notes:')
    data['cam_additional_notes'] = safe_str(ws.cell(add_notes_row, 4).value) if add_notes_row else None

    # ── TI ──
    ti_row, _ = find_label(ws, 'TI / SQ FT')
    if ti_row is None:
        ti_row, _ = find_label(ws, 'TI /SQ FT')
    if ti_row:
        data['ti_per_sf'] = safe_val(ws.cell(ti_row, 4).value)
        data['ti_total'] = get_value_right_of_label(ws, 'TI - Total:')
        # Comments below TI row
        ti_comments_row, _ = find_label(ws, 'Comments:', max_row=ti_row + 10)
        if ti_comments_row and ti_comments_row > ti_row:
            parts = [safe_str(ws.cell(ti_comments_row + r, 4).value) for r in range(0, 3)]
            data['ti_comments'] = ' '.join(filter(None, parts)) or None
        else:
            data['ti_comments'] = None
    else:
        data['ti_per_sf'] = None
        data['ti_total'] = None
        data['ti_comments'] = None

    # ── Insurance ──
    ins_row, _ = find_label(ws, 'Insurance Coverage:')
    if ins_row:
        data['insurance_type'] = safe_str(ws.cell(ins_row, 5).value)
        data['insurance_amount'] = get_value_right_of_label(ws, 'Covg. Amount:')
    else:
        data['insurance_type'] = None
        data['insurance_amount'] = None

    # ── Parking ──
    data['parking_total_stalls'] = get_value_right_of_label(ws, 'Total # of Stalls:')
    data['parking_reserved_stalls'] = get_value_right_of_label(ws, '# of Reserved Stalls:')
    data['parking_reserved_rate'] = get_value_right_of_label(ws, 'Reserved Rate:')
    data['parking_unreserved_stalls'] = get_value_right_of_label(ws, '# of Unreserved Stalls:')
    data['parking_unreserved_rate'] = get_value_right_of_label(ws, 'Unreserved Rate:')

    return data


def parse_format_b(ws):
    """Parse Format B abstract (515 Westheimer compact template)."""
    data = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        for cell in row:
            v = str(cell.value or '').strip().rstrip(':')
            if not v:
                continue
            val = safe_val(ws.cell(row=cell.row, column=cell.column + 2).value) if cell.column + 2 <= ws.max_column else None
            if v == 'TENANT DBA':
                data['tenant_dba'] = val
            elif v == 'TENANT':
                data['tenant_legal_name'] = val
            elif v == 'SQUARE FOOTAGE':
                data['tenant_sq_feet'] = val
            elif v == 'SPECIFIC USE':
                data['permitted_use'] = val
            elif v == 'RENT COMMENCEMENT DATE':
                data['rent_commencement'] = val
            elif v == 'LEASE EXPIRATION DATE':
                data['lease_expiration'] = val
            elif v == 'TERM':
                data['term_months'] = val
            elif v == 'SECURITY DEPOSIT':
                data['security_deposit'] = val
            elif v == 'GUARANTOR':
                data['guarantor'] = val
            elif v == 'EXCLUSIVES / RESTRICTIONS':
                data['exclusives'] = val
            elif v == 'RENEWAL OPTION(S)':
                data['renewal_options'] = val
            elif v == 'PERCENTAGE RENT':
                data['percentage_rent'] = {'percent': val}

    rent_schedule = []
    for r in range(14, 30):
        period = safe_str(ws.cell(r, 5).value)
        monthly = safe_val(ws.cell(r, 6).value)
        annual = safe_val(ws.cell(r, 7).value)
        psf = safe_val(ws.cell(r, 8).value)
        if period and monthly:
            rent_schedule.append({'period': period, 'monthly': monthly, 'annual': annual, 'psf': psf})
    data['rent_schedule'] = rent_schedule
    if rent_schedule:
        data['base_rent_monthly'] = rent_schedule[0]['monthly']
        data['base_rent_annual'] = rent_schedule[0]['annual']
        data['rent_per_sf_annual'] = rent_schedule[0]['psf']
    return data


def main():
    data_dir = os.path.join(os.path.dirname(__file__), '..', 'data')
    abstract_dir = os.path.join(data_dir, 'abstracts', 'Abstracts')

    with open(os.path.join(data_dir, 'lease_abstract_mapping.json')) as f:
        mapping = json.load(f)

    gold_standard = []
    errors = []

    for pair in mapping:
        try:
            path = os.path.join(abstract_dir, pair['abstract_file'])
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb[pair['abstract_sheet']] if pair['abstract_sheet'] else wb.active
            r4 = str(ws.cell(4, 2).value or ws.cell(4, 3).value or '')
            data = parse_format_a(ws) if 'PROPERTY' in r4 else parse_format_b(ws)
            wb.close()
            gold_standard.append({
                'lease_file': pair['lease_file'],
                'abstract_file': pair['abstract_file'],
                'abstract_sheet': pair['abstract_sheet'],
                'tenant': pair['tenant_dba'] or pair['tenant_legal'],
                'ground_truth': data
            })
        except Exception as e:
            errors.append(f"{pair['abstract_file']}: {e}")

    output_path = os.path.join(data_dir, 'gold_standard.json')
    with open(output_path, 'w') as f:
        json.dump(gold_standard, f, indent=2, default=str)

    print(f"Gold standard: {len(gold_standard)} entries")
    if errors:
        print(f"Errors ({len(errors)}):")
        for e in errors:
            print(f"  {e}")
    else:
        print("No errors.")

    # Field coverage
    fields = {}
    for e in gold_standard:
        for k, v in e['ground_truth'].items():
            if v is not None and v != '' and v != [] and v != {}:
                fields[k] = fields.get(k, 0) + 1

    print(f"\nField coverage ({len(gold_standard)} entries):")
    for k in sorted(fields, key=lambda x: -fields[x]):
        print(f"  {k:<28} {fields[k]:>3}/{len(gold_standard)} ({fields[k] / len(gold_standard) * 100:.0f}%)")


if __name__ == '__main__':
    main()
